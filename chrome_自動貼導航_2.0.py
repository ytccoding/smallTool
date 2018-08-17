# -*- coding:utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os ,ytFuntion
from time import sleep
from openpyxl import workbook , load_workbook

#更改寫法
#修改認證碼打錯後自動再填入帳密

wb = load_workbook("小工具設定.xlsx")
sheet = wb["帳號"] # 獲取一張表
for i in range(1,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "google":
      google_Account_sitting = str(sheet["C" + str(i)].value).strip()
      google_Password_sitting = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "總控":
      control_Account = str(sheet["C" + str(i)].value).strip()
      control_Password = str(sheet["D" + str(i)].value).strip()
      
sheet = wb["url"] # 獲取一張表
for i in range(1,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "總控":
      control_Url = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "總控站點":
      control_Sitting_url = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "APKIOS":
      google_Driver_url = str(sheet["D" + str(i)].value).strip()


if os.path.isfile("DEV-ipa%2Fapk 下載位置.xlsx"):    #先確認檔案是否存在
   os.remove("DEV-ipa%2Fapk 下載位置.xlsx")

options = webdriver.ChromeOptions()
prefs = {"profile.default_content_settings.popups": 0, "download.default_directory": os.getcwd()}
options.add_experimental_option("prefs", prefs)

control_Web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe', chrome_options=options))
control_Web.webDriver.get(google_Driver_url) #目標網址

control_Web.elementSendKeys("identifierId" ,1 ,text = google_Account_sitting)
control_Web.elementClick("div[id='identifierNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)

control_Web.elementSendKeys("input[type=password]" ,6 ,text = google_Password_sitting)
control_Web.elementClick("div[id='passwordNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)

sleep(5)
if not os.path.isfile("DEV-ipa%2Fapk 下載位置.xlsx"):
   input("確認google有成功登入後,請按ENTER繼續")

url_Number = " "
error_Count = 0
error_Count_s = 0
while url_Number[0].strip().upper() != "Q":
   wb = load_workbook("DEV-ipa%2Fapk 下載位置.xlsx",data_only=True) # 打開一個活頁薄
   wb.save(r"DEV-ipa%2Fapk 下載位置.xlsx") #Excel公式處理
   wb = load_workbook("DEV-ipa%2Fapk 下載位置.xlsx")
   sheet = wb["下載地點"] # 獲取一張表
   if url_Number == " ":
      url_Number = input("序列號:")
      aORi_Setting = input("APK或iOS:")
   url = ""
   url_name = ""
   if control_Web.webDriver.current_url != control_Sitting_url:
      keySecurityCode = "NG"
      while keySecurityCode == "NG":
         control_Web.webDriver.get(control_Url)
         control_Web.elementSendKeys("username" ,1 ,text = control_Account)
         control_Web.elementSendKeys("password" ,1 ,text = control_Password)
         control_Web.elementSendKeys("securityCode" ,1 ,text = input("認證碼:"))
         control_Web.elementClick("[class='btn btn-lg btn-warning btn-block']" ,6)
         control_Web.webDriver.get(control_Sitting_url)
         
         sleep(1)
         
         if control_Web.webDriver.current_url != control_Sitting_url:
            keySecurityCode = "NG"
         else:
            keySecurityCode = "OK"

      control_Web.webDriver.get(control_Sitting_url)
      control_Web.element("input[name=iosPath" ,6)

   for i in range(1 ,len(sheet["A"])+1):
      if sheet["A" + str(i)].value == url_Number.strip().upper():
         if url_Number[0].upper().strip() != "Q" and aORi_Setting[0].strip().lower() == "i":
            url = sheet["D" + str(i)].value
            url_name = sheet["B" + str(i)].value
            ios_Download = control_Web.webDriver.find_element_by_css_selector("input[name=iosPath]")
            ios_Download.clear()
            url = url.strip()
            for j in range(len(url)):
               ios_Download.send_keys(url[j])
            control_Web.elementClick("button[type=submit]" ,6)
            url_Number = " "
            error_Count = 0
            error_Count_s = 0
            print(url.strip())
            print(url_name.strip())
            print()
            break
         elif url_Number[0].upper().strip() != "Q" and  aORi_Setting[0].strip().lower() == "a":
            url = sheet["E" + str(i)].value
            url_name = sheet["B" + str(i)].value
            apk_Download = control_Web.webDriver.find_element_by_css_selector("input[name=androidPath]")
            apk_Download.clear()
            url = url.strip()
            for j in range(len(url)):
               apk_Download.send_keys(url[j])
            control_Web.elementClick("button[type=submit]" ,6)
            url_Number = " "
            error_Count = 0
            error_Count_s = 0
            print(url.strip())
            print(url_name.strip())
            print()
            break
      else:
         error_Count +=1
   if url_Number[0].upper().strip() != "Q" and error_Count != 0:
      print("無此站點,開始重新下載")
      if os.path.isfile("DEV-ipa%2Fapk 下載位置.xlsx"):    #先確認檔案是否存在
         os.remove("DEV-ipa%2Fapk 下載位置.xlsx")
      control_Web.webDriver.get(google_Driver_url)
      sleep(5)
      control_Web.webDriver.get(control_Sitting_url)
      error_Count_s += 1
      if error_Count_s == 2:
         print("站點獲取重試失敗")
         print()
         url_Number = " "
control_Web.webDriver.quit()
   

