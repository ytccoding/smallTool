# -*- coding:utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
from time import sleep
from openpyxl import workbook
from openpyxl import load_workbook
#配合GOOGLE及表單更新及部分寫法更改

wb = load_workbook("小工具設定.xlsx")
sheet = wb["帳號"]  # 獲取一張表
for i in range(1, len(sheet["B"]) + 1):
   if str(sheet["B" + str(i)].value).strip() == "google":
      google_Account_sitting = str(sheet["C" + str(i)].value).strip()
      google_Password_sitting = str(sheet["D" + str(i)].value).strip()
      
sheet = wb["url"]  # 獲取一張表
for i in range(1, len(sheet["B"]) + 1):
   if str(sheet["B" + str(i)].value).strip() == "商戶站點":
      google_Driver_url = str(sheet["D" + str(i)].value).strip()

if os.path.isfile("商戶站點地址.xlsx"):  # 先確認檔案是否存在
   os.remove("商戶站點地址.xlsx")

options = webdriver.ChromeOptions()
prefs = {"profile.default_content_settings.popups": 0, "download.default_directory": os.getcwd()}
options.add_experimental_option("prefs", prefs)

control_Web = webdriver.Chrome(executable_path='chromedriver.exe', chrome_options=options)
control_Web.get(google_Driver_url) #目標網址

WebDriverWait(control_Web, 10).until(EC.visibility_of_element_located((By.ID,"identifierId")))
google_Account = control_Web.find_element_by_id("identifierId")
google_Account.send_keys(google_Account_sitting)
btn_Submit = control_Web.find_element_by_css_selector("div[id='identifierNext'] content[class='CwaK9'] span[class='RveJvd snByac']")
btn_Submit.click()

WebDriverWait(control_Web, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"input[type=password]")))
google_Password = control_Web.find_element_by_css_selector("input[type=password]")
google_Password.send_keys(google_Password_sitting)
btn_Submit = control_Web.find_element_by_css_selector("div[id='passwordNext'] content[class='CwaK9'] span[class='RveJvd snByac']")
btn_Submit.click()
sleep(3)
if not os.path.isfile("商戶站點地址.xlsx"):
   input("確認google有成功登入後,請按ENTER繼續")

wb = load_workbook("商戶站點地址.xlsx",data_only=True) # 打開一個活頁薄
wb.save(r"商戶站點地址.xlsx") #Excel公式處理
wb = load_workbook("商戶站點地址.xlsx")

sheet = wb["站點資料"] # 獲取一張表
url_Number = " "

while url_Number[0].strip().upper() != "Q":
    url_Number = ""
    url_Number = list(input("序列號(用空白隔開):").split())
    url = []

    for i in range(1,len(sheet["D"])):
        for j in range(len(url_Number)):
            if sheet["D" + str(i)].value == url_Number[j].upper():
                url.append(sheet["F" + str(i)].value)
                control_Web.execute_script("window.open('http://" + str(sheet["F" + str(i)].value) + "')")
                
    handles = control_Web.window_handles
    sleep(5)
    for newhandle in handles:
         control_Web.switch_to_window(newhandle)
         try:
            WebDriverWait(control_Web, 10).until(EC.visibility_of_element_located((By.LINK_TEXT ,"手机购彩")))
            control_Web.find_element_by_link_text("手机购彩").click()
            sleep(1)
         except:
            pass
    print(url_Number)
    print(url)
    url_Number = " "

   

